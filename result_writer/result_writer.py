import openpyxl


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    ws.title = "Summary"
    return wb


def save_workbook(wb):
    wb.save("./resl.xlsx")


def create_new_sheet(wb, test_no, now):
    wb.create_sheet(test_no)
    ws = wb[test_no]
    ws['A1'] = f'テスト番号：{test_no}'
    ws['A2'] = f'実行時刻：{now}'
    ws['A3'] = "---------------------------------"

    return ws


def write_status_code(ws, expected_status_code, actual_status_code, count):
    ws['A' + str(count)] = str(expected_status_code)
    ws['B' + str(count)] = str(actual_status_code)
    if expected_status_code == actual_status_code:
        ws['C' + str(count)] = "OK"
    else:
        ws['C' + str(count)] = "NG"

    return ws, count + 1


def write_expect_actual(ws, count, title):
    # ws = wb[test_no]
    print('A' + str(count))
    ws['A' + str(count)] = title
    ws['A' + str(count + 1)] = "期待値"
    ws['B' + str(count)] = "実際値"
    ws['C' + str(count)] = "結果"
    ws['A' + str(count + 2)] = "---------------------------------"

    return ws, count + 3


def main():
    wb = create_workbook()
    create_new_sheet(wb)
    save_workbook(wb)


if __name__ == "__main__":
    main()
